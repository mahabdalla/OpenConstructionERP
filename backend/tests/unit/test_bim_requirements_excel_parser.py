"""Unit tests for the Excel/CSV generic parser."""

import tempfile
from pathlib import Path

import pytest

from app.modules.bim_requirements.parsers.excel_parser import (
    ExcelCSVParser,
    _parse_constraint_value,
)


@pytest.fixture
def parser() -> ExcelCSVParser:
    return ExcelCSVParser()


def _create_xlsx(headers: list[str], rows: list[list]) -> Path:
    """Create a temporary .xlsx file with given headers and rows."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = h  # type: ignore[union-attr]
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = val  # type: ignore[union-attr]

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
    return Path(tmp.name)


class TestConstraintValueParsing:
    """Tests for _parse_constraint_value."""

    def test_enum_semicolons(self) -> None:
        result = _parse_constraint_value("REI60; REI90; REI120")
        assert result["enum"] == ["REI60", "REI90", "REI120"]

    def test_range(self) -> None:
        result = _parse_constraint_value("0.1 - 1.0")
        assert result["min"] == 0.1
        assert result["max"] == 1.0

    def test_gte(self) -> None:
        result = _parse_constraint_value(">= 2.5 m")
        assert result["min"] == 2.5
        assert result["unit"] == "m"

    def test_lte(self) -> None:
        result = _parse_constraint_value("<= 100")
        assert result["max"] == 100.0

    def test_regex_pattern(self) -> None:
        result = _parse_constraint_value("^[A-Z]{2}")
        assert result["pattern"] == "^[A-Z]{2}"

    def test_datatype_keyword(self) -> None:
        result = _parse_constraint_value("TEXT")
        assert result["datatype"] == "IFCTEXT"

    def test_cardinality_keyword(self) -> None:
        result = _parse_constraint_value("required")
        assert result["cardinality"] == "required"

    def test_simple_value(self) -> None:
        result = _parse_constraint_value("F90")
        assert result["value"] == "F90"

    def test_empty_string(self) -> None:
        result = _parse_constraint_value("")
        assert result == {}


class TestExcelParser:
    """Tests for the Excel/CSV generic parser."""

    def test_parse_xlsx_with_standard_headers(self, parser: ExcelCSVParser) -> None:
        """Parse an Excel file with standard BIM requirement headers."""
        headers = ["IFC Class", "Property Set", "Property Name", "Value", "Unit"]
        rows = [
            ["IfcWall", "Pset_WallCommon", "FireRating", "REI60; REI90", ""],
            ["IfcDoor", "Pset_DoorCommon", "IsExternal", "required", ""],
        ]
        path = _create_xlsx(headers, rows)
        try:
            result = parser.parse(path)
            assert result.success
            assert len(result.requirements) == 2

            wall_req = result.requirements[0]
            assert wall_req.element_filter.get("ifc_class") == "IFCWALL"
            assert wall_req.property_group == "Pset_WallCommon"
            assert wall_req.property_name == "FireRating"
        finally:
            path.unlink(missing_ok=True)

    def test_parse_csv_string(self, parser: ExcelCSVParser) -> None:
        """Parse CSV content from a string."""
        csv_content = "Element,Property,Value\nIFCWALL,FireRating,REI60\n"
        result = parser.parse(csv_content)
        assert result.success
        assert len(result.requirements) == 1
        assert result.requirements[0].property_name == "FireRating"

    def test_empty_file(self, parser: ExcelCSVParser) -> None:
        """Empty file produces an error."""
        result = parser.parse("")
        assert not result.success

    def test_headers_only(self, parser: ExcelCSVParser) -> None:
        """File with headers but no data produces a warning."""
        path = _create_xlsx(["Property", "Value"], [])
        try:
            result = parser.parse(path)
            assert not result.success
            assert len(result.warnings) > 0
        finally:
            path.unlink(missing_ok=True)

    def test_column_auto_mapping(self, parser: ExcelCSVParser) -> None:
        """German headers are auto-mapped correctly."""
        headers = ["Bauteil", "Merkmalsgruppe", "Merkmal", "Wert"]
        rows = [["IfcBeam", "Pset_BeamCommon", "LoadBearing", "true"]]
        path = _create_xlsx(headers, rows)
        try:
            result = parser.parse(path)
            assert result.success
            req = result.requirements[0]
            assert req.element_filter.get("ifc_class") == "IFCBEAM"
            assert req.property_name == "LoadBearing"
        finally:
            path.unlink(missing_ok=True)

    def test_metadata_includes_mapping(self, parser: ExcelCSVParser) -> None:
        """Parse result metadata includes the column mapping."""
        csv_content = "Property,Value\nFireRating,REI60\n"
        result = parser.parse(csv_content)
        assert "column_mapping" in result.metadata
