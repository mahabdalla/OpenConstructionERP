"""Unit tests for the format classifier."""

import tempfile
from pathlib import Path

import pytest

from app.modules.bim_requirements.classifier import FormatClassifier


@pytest.fixture
def classifier() -> FormatClassifier:
    return FormatClassifier()


def _write_temp(content: str, suffix: str) -> Path:
    """Write content to a temp file with given suffix."""
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False, mode="w", encoding="utf-8") as tmp:
        tmp.write(content)
    return Path(tmp.name)


class TestFormatClassifier:
    """Tests for the format classifier."""

    def test_ids_xml(self, classifier: FormatClassifier) -> None:
        """IDS XML is detected from namespace."""
        content = """\
<?xml version="1.0"?>
<ids xmlns="http://standards.buildingsmart.org/IDS">
  <specifications/>
</ids>
"""
        path = _write_temp(content, ".ids")
        try:
            assert classifier.classify(path) == "IDS"
        finally:
            path.unlink()

    def test_ids_with_xml_extension(self, classifier: FormatClassifier) -> None:
        """IDS content with .xml extension is still detected as IDS."""
        content = '<ids xmlns="http://standards.buildingsmart.org/IDS"><specifications/></ids>'
        path = _write_temp(content, ".xml")
        try:
            assert classifier.classify(path) == "IDS"
        finally:
            path.unlink()

    def test_generic_xml(self, classifier: FormatClassifier) -> None:
        """Non-IDS XML is classified as GenericXML."""
        content = '<?xml version="1.0"?><root><data/></root>'
        path = _write_temp(content, ".xml")
        try:
            assert classifier.classify(path) == "GenericXML"
        finally:
            path.unlink()

    def test_csv_file(self, classifier: FormatClassifier) -> None:
        """CSV is always classified as Excel (generic parser)."""
        path = _write_temp("Property,Value\nFireRating,REI60\n", ".csv")
        try:
            assert classifier.classify(path) == "Excel"
        finally:
            path.unlink()

    def test_revit_sp_txt(self, classifier: FormatClassifier) -> None:
        """Revit Shared Parameters .txt is detected from header."""
        content = "# This is a Revit shared parameter file.\n*META\tVERSION\n"
        path = _write_temp(content, ".txt")
        try:
            assert classifier.classify(path) == "RevitSP"
        finally:
            path.unlink()

    def test_plain_txt(self, classifier: FormatClassifier) -> None:
        """Generic text file is classified as PlainText."""
        path = _write_temp("just some text", ".txt")
        try:
            assert classifier.classify(path) == "PlainText"
        finally:
            path.unlink()

    def test_bimq_json(self, classifier: FormatClassifier) -> None:
        """BIMQ JSON is detected from structure."""
        content = '{"concept_tree": {"elements": []}}'
        path = _write_temp(content, ".json")
        try:
            assert classifier.classify(path) == "BIMQ"
        finally:
            path.unlink()

    def test_unsupported_extension(self, classifier: FormatClassifier) -> None:
        """Unsupported extension raises ValueError."""
        path = _write_temp("data", ".xyz")
        try:
            with pytest.raises(ValueError, match="Unsupported file format"):
                classifier.classify(path)
        finally:
            path.unlink()

    def test_cobie_xlsx(self, classifier: FormatClassifier) -> None:
        """COBie Excel is detected from sheet names."""
        import openpyxl

        wb = openpyxl.Workbook()
        wb.create_sheet("Component")
        wb.create_sheet("Type")
        wb.create_sheet("Attribute")
        wb.create_sheet("Floor")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
        path = Path(tmp.name)
        try:
            assert classifier.classify(path) == "COBie"
        finally:
            path.unlink()

    def test_generic_xlsx(self, classifier: FormatClassifier) -> None:
        """Regular Excel file is classified as Excel."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "Property"  # type: ignore[union-attr]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
        path = Path(tmp.name)
        try:
            assert classifier.classify(path) == "Excel"
        finally:
            path.unlink()
